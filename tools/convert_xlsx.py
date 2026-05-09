from __future__ import annotations

import json
import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from datetime import date, datetime, timedelta
from pathlib import Path


NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def read_xml(archive: zipfile.ZipFile, name: str) -> ET.Element:
    return ET.fromstring(archive.read(name))


def shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []

    root = read_xml(archive, "xl/sharedStrings.xml")
    values: list[str] = []
    for item in root.findall("main:si", NS):
        values.append("".join(text.text or "" for text in item.findall(".//main:t", NS)))
    return values


def sheet_paths(archive: zipfile.ZipFile) -> list[tuple[str, str]]:
    workbook = read_xml(archive, "xl/workbook.xml")
    rels = read_xml(archive, "xl/_rels/workbook.xml.rels")
    rel_targets = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels.findall("pkgrel:Relationship", NS)
    }

    paths: list[tuple[str, str]] = []
    for sheet in workbook.findall("main:sheets/main:sheet", NS):
        sheet_name = sheet.attrib["name"]
        rel_id = sheet.attrib[f"{{{NS['rel']}}}id"]
        target = rel_targets[rel_id].lstrip("/")
        if not target.startswith("xl/"):
            target = f"xl/{target}"
        paths.append((sheet_name, target))
    return paths


def cell_value(cell: ET.Element, strings: list[str]) -> str:
    value = cell.find("main:v", NS)
    if value is None:
        inline = cell.find("main:is", NS)
        if inline is None:
            return ""
        return "".join(text.text or "" for text in inline.findall(".//main:t", NS))

    raw = value.text or ""
    if cell.attrib.get("t") == "s":
        return strings[int(raw)] if raw.isdigit() and int(raw) < len(strings) else raw
    return raw


def rows_from_sheet(archive: zipfile.ZipFile, sheet_path: str, strings: list[str]) -> list[list[str]]:
    root = read_xml(archive, sheet_path)
    rows: list[list[str]] = []
    for row in root.findall("main:sheetData/main:row", NS):
        values = [cell_value(cell, strings).strip() for cell in row.findall("main:c", NS)]
        rows.append(values)
    return rows


def excel_date(value: str) -> str:
    if re.fullmatch(r"\d+(\.\d+)?", value):
        date = datetime(1899, 12, 30) + timedelta(days=float(value))
        return date.date().isoformat()

    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(value, pattern).date().isoformat()
        except ValueError:
            pass

    return ""


def normalize_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return excel_date(str(value).strip())


def parse_money(value: str) -> float:
    cleaned = value.replace(",", "").replace("$", "").strip()
    return round(float(cleaned), 2)


def find_header_start(row: tuple[object, ...]) -> int | None:
    values = [str(cell).strip() if cell is not None else "" for cell in row]
    target = ["日期", "東西", "錢", "備注"]
    for index in range(0, max(len(values) - len(target) + 1, 0)):
        if values[index : index + len(target)] == target:
            return index
    return None


def convert_with_openpyxl(path: Path) -> list[dict[str, object]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    transactions: list[dict[str, object]] = []

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        for row_index, row in enumerate(rows):
            header_start = find_header_start(row)
            if header_start is None:
                continue

            for data_index, data_row in enumerate(rows[row_index + 1 :], start=1):
                if not any(cell is not None and str(cell).strip() for cell in data_row):
                    continue

                raw_date = data_row[header_start] if len(data_row) > header_start else ""
                raw_description = data_row[header_start + 1] if len(data_row) > header_start + 1 else ""
                raw_money = data_row[header_start + 2] if len(data_row) > header_start + 2 else ""
                raw_note = data_row[header_start + 3] if len(data_row) > header_start + 3 else ""
                description = str(raw_description).strip() if raw_description is not None else ""
                note = str(raw_note).strip() if raw_note is not None else ""

                parsed_date = normalize_date(raw_date)
                if isinstance(raw_money, (int, float)):
                    amount = round(float(raw_money), 2)
                else:
                    try:
                        amount = parse_money(str(raw_money))
                    except ValueError:
                        continue

                if not parsed_date or not description or amount == 0:
                    continue

                transactions.append(
                    {
                        "id": f"{worksheet.title}-{parsed_date}-{data_index}",
                        "date": parsed_date,
                        "type": "expense" if amount >= 0 else "income",
                        "category": description,
                        "note": note,
                        "amount": abs(amount),
                    }
                )
            break

    return transactions


def convert(path: Path) -> list[dict[str, object]]:
    transactions: list[dict[str, object]] = []

    with zipfile.ZipFile(path) as archive:
        strings = shared_strings(archive)
        for sheet_name, sheet_path in sheet_paths(archive):
            rows = rows_from_sheet(archive, sheet_path, strings)
            for index, row in enumerate(rows):
                normalized = [cell.strip() for cell in row]
                if normalized[:4] == ["日期", "東西", "錢", "備注"]:
                    data_rows = rows[index + 1 :]
                    for data_index, data_row in enumerate(data_rows, start=1):
                        if len(data_row) < 3 or not any(data_row):
                            continue

                        date = excel_date(data_row[0])
                        description = data_row[1].strip() if len(data_row) > 1 else ""
                        money = data_row[2].strip() if len(data_row) > 2 else ""
                        note = data_row[3].strip() if len(data_row) > 3 else ""

                        if not date or not description or not money:
                            continue

                        try:
                            amount = parse_money(money)
                        except ValueError:
                            continue

                        transactions.append(
                            {
                                "id": f"{sheet_name}-{date}-{data_index}",
                                "date": date,
                                "type": "expense" if amount >= 0 else "income",
                                "category": description,
                                "note": note,
                                "amount": abs(amount),
                            }
                        )
                    break

    return transactions


if __name__ == "__main__":
    source = Path(sys.argv[1])
    destination = Path(sys.argv[2])
    destination.parent.mkdir(parents=True, exist_ok=True)
    try:
        transactions = convert_with_openpyxl(source)
        parser = "openpyxl"
    except ModuleNotFoundError:
        transactions = convert(source)
        parser = "xml fallback"

    destination.write_text(
        json.dumps(transactions, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"Wrote {len(transactions)} transactions to {destination} using {parser}")
